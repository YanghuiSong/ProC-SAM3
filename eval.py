import os
import os.path as osp
import argparse
import openpyxl
import pickle
from mmengine.runner import Runner
from mmengine.config import Config, DictAction
import random
import numpy as np
from torch.utils.data import Subset, Sampler
import torch
import re
from difflib import SequenceMatcher

import sam3_segmentor as sam3_segmentor
import sam3_segmentor_cached as sam3_segmentor_cached  # Import the cached version

import custom_datasets

from core.qwen_agent import QwenAgent



def parse_args():
    parser = argparse.ArgumentParser(
        description='CorrCLIP evaluation with MMSeg')
    parser.add_argument('config', default='./configs/cfg_loveda.py')
    parser.add_argument(
        '--show', action='store_true', help='show prediction results')
    parser.add_argument(
        '--show_dir',
        default='./show_dir/',
        help='directory to save visualizaion images')
    parser.add_argument(
        '--out',
        type=str,
        help='The directory to save output prediction for offline evaluation')
    parser.add_argument(
        '--percent',
        type=float,
        default=100.0,
        help='Percentage of data to use for evaluation (default: 100%%)')
    parser.add_argument(
        '--seed',
        type=int,
        default=42,
        help='Random seed for data sampling (default: 42)')
    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction,
        help='override some settings in the used config, the key-value pair '
        'in xxx=yyy format will be merged into config file. If the value to '
        'be overwritten is a list, it should be like key="[a,b]" or key=a,b '
        'It also allows nested list/tuple values, e.g. key="[(a,b),(c,d)]" '
        'Note that the quotation marks are necessary and that no white space '
        'is allowed.')
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none',
        help='job launcher')
    # Support both --local_rank and --local-rank for PyTorch launcher compatibility.
    parser.add_argument('--local_rank', '--local-rank', type=int, default=0)
    args = parser.parse_args()
    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)

    return args


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'
        sheet['F1'] = 'AP50'
        sheet['G1'] = 'AP75'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['Model'])
        sheet.cell(row=last_row + index, column=2, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=3, value=result.get('aAcc', 'N/A'))
        sheet.cell(row=last_row + index, column=4, value=result.get('mIoU', 'N/A'))
        sheet.cell(row=last_row + index, column=5, value=result.get('mAcc', 'N/A'))
        sheet.cell(row=last_row + index, column=6, value=result.get('AP50', 'N/A'))
        sheet.cell(row=last_row + index, column=7, value=result.get('AP75', 'N/A'))

    workbook.save(file_path)


def trigger_visualization_hook(cfg, args):
    default_hooks = cfg.default_hooks
    if 'visualization' in default_hooks:
        visualization_hook = default_hooks['visualization']
        # Turn on visualization
        visualization_hook['draw'] = True
        if args.show:
            visualization_hook['show'] = True
            visualization_hook['wait_time'] = args.wait_time
        if args.show_dir:
            visualizer = cfg.visualizer
            visualizer['save_dir'] = args.show_dir
    else:
        raise RuntimeError(
            'VisualizationHook must be included in default_hooks.'
            'refer to usage '
            '"visualization=dict(\'VisualizationHook\')"')

    return cfg


class PercentSampler(Sampler):
    """A sampler that samples a percentage of the dataset"""
    def __init__(self, dataset, percent, seed=42):
        self.dataset = dataset
        self.percent = percent
        self.seed = seed
        
        if percent >= 100.0:
            self.indices = list(range(len(dataset)))
        else:
            # Set seed for reproducibility
            rng = random.Random(seed)
            n_samples = max(1, int(len(dataset) * percent / 100.0))
            self.indices = rng.sample(range(len(dataset)), n_samples)
            self.indices.sort()  # Sort for consistency

    def __iter__(self):
        return iter(self.indices)

    def __len__(self):
        return len(self.indices)


def modify_runner_for_percent_data(runner, percent, seed):
    """
    Modify the runner to use only a percentage of the data
    """
    if percent >= 100.0:
        return runner
    
    # Set random seed for reproducibility
    random.seed(seed)
    np.random.seed(seed)
    
    # Get the test dataset
    test_dataset = runner.test_loop.dataloader.dataset
    
    # Create a subset sampler
    subset_sampler = PercentSampler(test_dataset, percent, seed)
    
    # Get the original dataloader's parameters
    orig_dataloader = runner.test_loop.dataloader
    
    # Create a new dataloader with the subset sampler
    from torch.utils.data import DataLoader
    new_dataloader = DataLoader(
        dataset=test_dataset,
        batch_size=orig_dataloader.batch_size,
        sampler=subset_sampler,
        num_workers=orig_dataloader.num_workers,
        collate_fn=orig_dataloader.collate_fn,
        pin_memory=orig_dataloader.pin_memory,
        drop_last=False,  # Don't drop last batch
        timeout=orig_dataloader.timeout,
        persistent_workers=orig_dataloader.persistent_workers,
    )
    
    # Replace the dataloader in the test loop
    runner.test_loop.dataloader = new_dataloader
    
    print(f"Using {len(subset_sampler)}/{len(test_dataset)} samples ({percent}%) for evaluation")
    
    return runner


def normalize_optimize_method(optimize_method):
    method = str(optimize_method).strip().lower()
    alias_map = {
        'preset': {'preset'},
        'guided': {'guided'},
        'adaptive': {'adaptive'},
        'hybrid': {'hybrid'},
        'hybrid_strict': {
            'hybrid_strict',
            'hybrid_str',
            'hybid_strict',
            'hybid_str',
            'hybrid-strict',
            'hybrid strict',
        },
    }

    for canonical_name, aliases in alias_map.items():
        if method in aliases:
            if method != canonical_name:
                print(
                    f"Normalizing optimize_method from '{optimize_method}' "
                    f"to '{canonical_name}'."
                )
            return canonical_name

    valid_methods = ", ".join(sorted(alias_map))
    raise ValueError(
        f"Unsupported optimize_method '{optimize_method}'. "
        f"Expected one of: {valid_methods}."
    )


def normalize_prompt_token(token):
    token = str(token).strip()
    token = token.replace('锛?, ', '').replace('_', ' ').replace('-', ' ')
    token = re.sub(r'\s+', ' ', token)
    return token.lower().strip(' ,')


def match_base_class_name(candidate, base_classes):
    normalized_candidate = normalize_prompt_token(candidate)
    if not normalized_candidate:
        return candidate

    normalized_lookup = {
        normalize_prompt_token(base_class): base_class for base_class in base_classes
    }
    if normalized_candidate in normalized_lookup:
        return normalized_lookup[normalized_candidate]

    candidate_tokens = set(normalized_candidate.split())
    best_match = candidate
    best_score = 0.0
    best_overlap = 0.0
    best_seq_ratio = 0.0
    for base_class in base_classes:
        normalized_base = normalize_prompt_token(base_class)
        base_tokens = set(normalized_base.split())
        token_overlap = len(candidate_tokens & base_tokens) / max(
            len(candidate_tokens | base_tokens), 1
        )
        seq_ratio = SequenceMatcher(
            None, normalized_candidate, normalized_base
        ).ratio()
        score = 0.65 * seq_ratio + 0.35 * token_overlap
        if score > best_score:
            best_match = base_class
            best_score = score
            best_overlap = token_overlap
            best_seq_ratio = seq_ratio

    if best_score >= 0.55:
        return best_match
    if best_overlap > 0.0 and best_seq_ratio >= 0.3:
        return best_match
    return candidate


def prompt_pool_matches_exp_file(prompt_pool_path, exp_classname_path):
    try:
        with open(prompt_pool_path, 'rb') as f:
            prompt_pool = pickle.load(f)
    except Exception:
        return False

    if not isinstance(prompt_pool, dict):
        return False

    if not exp_classname_path or not os.path.exists(exp_classname_path):
        return False

    expected_prompt_pool = load_expanded_prompts_from_exp_file(exp_classname_path)
    return prompt_pool == expected_prompt_pool


def exp_file_has_additional_prompts(exp_classname_path):
    try:
        with open(exp_classname_path, 'r', encoding='utf-8') as f:
            for line in f:
                tokens = [
                    part.strip()
                    for part in line.replace('锛?, ', '').split(',')
                    if part.strip()
                ]
                if len(tokens) > 1:
                    return True
    except OSError:
        return False
    return False


def prompt_pool_contains_expansions(prompt_pool_path):
    try:
        with open(prompt_pool_path, 'rb') as f:
            prompt_pool = pickle.load(f)
    except Exception:
        return False

    if not isinstance(prompt_pool, dict):
        return False

    for prompt_variants in prompt_pool.values():
        if isinstance(prompt_variants, (list, tuple)) and len(prompt_variants) > 1:
            return True
    return False


def check_and_generate_expanded_prompt_pool(cfg, args):
    """
    Ensure the expanded prompt pool exists when the selected model requires it.
    """
    # Get the model config.
    model_cfg = cfg.get('model', {})
    
    # Check whether this model type uses an expanded prompt pool.
    model_type = model_cfg.get('type', '')
    if (
        'SegEarthOV3Segmentation' not in model_type
        and 'CachedSegEarthOV3Segmentation' not in model_type
        and 'CachedSAM3OpenSegmentor' not in model_type
    ):
        print("Model does not require expanded prompt pool, skipping check.")
        return
    
    # Check whether expanded prompts are enabled in the model config.
    enable_expanded_prompt = model_cfg.get('enable_expanded_prompt', False)
    if not enable_expanded_prompt:
        print("Expanded prompt pool is not enabled in model config, skipping check.")
        return
    
    # Build the output filename from the config name and optimization method.
    config_basename = os.path.basename(args.config).replace('.py', '')
    
    # Resolve the optimization method.
    optimize_method = normalize_optimize_method(
        model_cfg.get('optimize_method', 'guided')
    )
    model_cfg['optimize_method'] = optimize_method
    cfg.model = model_cfg
    
    output_path = f"./prompt_pools/{config_basename}_expanded_prompt_pool_{optimize_method}.pkl"
    
    # Ensure the output directory exists.
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    classname_path = model_cfg.get('classname_path', '')
    exp_classname_path = classname_path.replace('.txt', '_exp.txt') if classname_path else ''
    should_regenerate = not os.path.exists(output_path)
    if not should_regenerate:
        source_paths = [path for path in [classname_path, exp_classname_path] if path and os.path.exists(path)]
        if source_paths:
            output_mtime = os.path.getmtime(output_path)
            latest_source_mtime = max(os.path.getmtime(path) for path in source_paths)
            if latest_source_mtime > output_mtime:
                print(f"Expanded prompt pool source updated after {output_path}, regenerating...")
                should_regenerate = True
    if (
        not should_regenerate
        and optimize_method == 'preset'
        and exp_classname_path
        and os.path.exists(exp_classname_path)
        and exp_file_has_additional_prompts(exp_classname_path)
        and not prompt_pool_contains_expansions(output_path)
    ):
        print(
            f"Existing expanded prompt pool at {output_path} does not contain "
            "the extra prompts declared in the exp file, regenerating..."
        )
        should_regenerate = True
    if (
        not should_regenerate
        and optimize_method == 'hybrid_strict'
        and output_path
        and os.path.exists(output_path)
        and exp_classname_path
        and os.path.exists(exp_classname_path)
        and not prompt_pool_matches_exp_file(output_path, exp_classname_path)
    ):
        print(
            f"Existing hybrid_strict prompt pool at {output_path} contains "
            "prompts that do not exactly match the exp file, regenerating..."
        )
        should_regenerate = True
    
    # Generate the expanded prompt pool when it is missing or stale.
    if should_regenerate:
        print(f"Generating expanded prompt pool at {output_path}...")
        
        # Always prefer classname_path so generated prompts stay aligned with model classes.
        if classname_path and os.path.exists(classname_path):
            with open(classname_path, 'r') as f:
                class_names = [line.strip() for line in f.readlines() if line.strip()]
        else:
            # Fall back to config class_names when classname_path is unavailable.
            class_names = cfg.get('class_names', [])
        
        class_names = [class_name.split(',')[0].strip() for class_name in class_names]

        if not class_names:
            print("Error: class_names not found in config, cannot generate expanded prompt pool")
            return
        
        # Get the dataset type for QwenAgent.
        dataset_type = cfg.get('dataset_type', '').lower()
        
        # Handle preset mode separately.
        if optimize_method == 'preset':
            print("Using preset optimization method. Loading predefined prompts.")
            
            # Check whether the exp file is available.
            if os.path.exists(exp_classname_path):
                print(f"Exp file found: {exp_classname_path}, using it as constraint for Qwen3 inference")
            else:
                print(f"Exp file not found: {exp_classname_path}, will use base classes only")
                
            # Initialize QwenAgent with the dataset name.
            qwen_agent = QwenAgent(dataset_name=dataset_type)
            
            # Generate expanded prompts under preset-mode constraints.
            expanded_prompts = qwen_agent.generate_expanded_class_prompts(
                base_classes=class_names,
                image_paths=[],  # Preset mode does not require sample images.
                output_path=output_path,
                optimize_method=optimize_method
            )
            
            # Store the prompt-pool path before returning from preset mode.
            model_cfg['expanded_prompt_pool_path'] = output_path
            cfg.model = model_cfg
            return

        if optimize_method == 'hybrid_strict':
            print("Using hybrid_strict optimization method. Refining prompts under exp-file constraints.")
            predefined_prompts = None
            if os.path.exists(exp_classname_path):
                print(f"Exp file found: {exp_classname_path}, using it as strict constraint set")
                predefined_prompts = load_expanded_prompts_from_exp_file(exp_classname_path)
            else:
                print(f"Exp file not found: {exp_classname_path}, falling back to adaptive prompts")

        # Non-preset modes need the dataset root and sample images.
        data_root = cfg.get('data_root', '')
        if not data_root:
            print("Error: data_root not found in config, cannot generate expanded prompt pool")
            return
        
        # Collect candidate validation image directories.
        candidate_img_dirs = []
        test_dataset_cfg = cfg.get('test_dataloader', {}).get('dataset', {})
        data_prefix = test_dataset_cfg.get('data_prefix', {})
        test_img_dir = data_prefix.get('img_path')
        if test_img_dir:
            candidate_img_dirs.append(
                test_img_dir if os.path.isabs(test_img_dir) else os.path.join(data_root, test_img_dir)
            )

        legacy_val_cfg = cfg.get('data', {}).get('val', {})
        legacy_img_dir = legacy_val_cfg.get('img_dir')
        if legacy_img_dir:
            candidate_img_dirs.append(
                legacy_img_dir if os.path.isabs(legacy_img_dir) else os.path.join(data_root, legacy_img_dir)
            )

        candidate_img_dirs.extend([
            os.path.join(data_root, 'img_dir', 'val'),
            os.path.join(data_root, 'img_dir', 'train'),
            os.path.join(data_root, 'val', 'src'),
            os.path.join(data_root, 'test', 'src'),
        ])
        val_img_dir = next((path for path in candidate_img_dirs if path and os.path.exists(path)), '')
        if not os.path.exists(val_img_dir):
            # Try the training directory if validation images are unavailable.
            val_img_dir = os.path.join(data_root, 'img_dir', 'train')
            if not os.path.exists(val_img_dir):
                print(
                    f"Warning: No valid image directory found in {data_root}. "
                    "Proceeding without sample images for prompt expansion."
                )
                val_img_dir = ''
        
        # Collect candidate image files for prompt expansion.
        image_extensions = ['.jpg', '.jpeg', '.png', '.tif', '.tiff']
        image_paths = []
        
        if val_img_dir:
            for root, dirs, files in os.walk(val_img_dir):
                for file in files:
                    if any(file.lower().endswith(ext) for ext in image_extensions):
                        image_paths.append(os.path.join(root, file))
        
        # Limit the number of sample images to keep prompt generation fast.
        if len(image_paths) > 5:
            image_paths = image_paths[:5]
        
        print(f"Found {len(image_paths)} images to process for prompt expansion (limit: 5)")
        print(f"Classes: {class_names}")
        
        # Only guided and hybrid modes preload prompts from the class file.
        if optimize_method != 'hybrid_strict':
            predefined_prompts = None
        if optimize_method in ['guided', 'hybrid']:
            # Load prompt hints from the main class file.
            if os.path.exists(classname_path):
                predefined_prompts = load_best_prompts_from_main_classfile(classname_path)
                print(f"Loaded best prompts from main classfile: {classname_path}")
        
        # Initialize QwenAgent with the dataset name.
        qwen_agent = QwenAgent(dataset_name=dataset_type)
        
        # Generate the expanded prompt pool.
        expanded_prompts = qwen_agent.generate_expanded_class_prompts(
            base_classes=class_names,
            image_paths=image_paths,
            output_path=output_path,
            predefined_prompts=predefined_prompts,
            optimize_method=optimize_method
        )
        
        print(f"Expanded prompt pool generation completed. Saved to {output_path}")
    else:
        print(f"Using existing expanded prompt pool at {output_path}")
        # Uncomment these lines to force regeneration during debugging.
        # os.remove(output_path)
        # print(f"Existing file deleted. Please re-run to regenerate.")
    
    # Keep the model config updated with the expanded prompt-pool path.
    # This applies whether the pool was generated or reused.
    model_cfg['expanded_prompt_pool_path'] = output_path
    cfg.model = model_cfg


def load_expanded_prompts_from_exp_file(exp_classname_path):
    """
    Load prompt expansions from the exp file literally for hybrid_strict mode.
    """
    def normalize_exp_token(token: str) -> str:
        token = token.strip()
        token = token.replace('锛?, ', '').replace('閿?, ', '')
        token = token.replace('瀛╝tural', 'natural')
        token = token.replace('瀵塭llow', 'yellow')
        token = re.sub(r'\s+', ' ', token).strip(' ,')
        return token

    predefined_prompts = {}
    with open(exp_classname_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    for line in lines:
        line = normalize_exp_token(line)
        if not line:
            continue

        parts = [
            normalize_exp_token(part)
            for part in line.split(',')
            if normalize_exp_token(part)
        ]
        if not parts:
            continue

        class_name = parts[0]
        literal_prompts = []
        for part in parts:
            if part not in literal_prompts:
                literal_prompts.append(part)
        predefined_prompts[class_name] = literal_prompts

    return predefined_prompts

def load_best_prompts_from_main_classfile(classname_path):
    """
    Load prompt variants from the main class-name file.
    """
    predefined_prompts = {}
    
    with open(classname_path, 'r') as f:
        lines = f.readlines()
    
    for line in lines:
        line = line.strip()
        if ',' in line:
            parts = line.split(',')
            class_name = parts[0].strip()
            # Include all comma-separated prompt variants from the class file.
            variations = [part.strip() for part in parts if part.strip() and part.strip() != class_name]
            predefined_prompts[class_name] = [class_name] + variations
        else:
            predefined_prompts[line] = [line]
    
    return predefined_prompts


def check_and_generate_prompt_pool(cfg):
    """
    Ensure the standard prompt pool exists when the selected model requires it.
    """
    # Get the model config.
    model_cfg = cfg.get('model', {})
    
    # Check whether this model type requires a standard prompt pool.
    model_type = model_cfg.get('type', '')
    if 'DynamicPromptEnhancedSegEarthOV3' not in model_type and 'AutoPromptEnhancedSegEarth' not in model_type:
        print("Model does not require prompt pool, skipping check.")
        return
    
    # Resolve the target prompt pool path.
    prompt_pool_path = model_cfg.get('prompt_pool_path', './prompt_pool.pkl')
    pool_building_percent = model_cfg.get('pool_building_percent', 10.0)
    
    # Generate the prompt pool when it is missing.
    if not os.path.exists(prompt_pool_path):
        print(f"Prompt pool not found at {prompt_pool_path}, generating now...")
        
        # Import the prompt-pool generation helper.
        from generate_prompt_pool import generate_prompt_pool
        
        # Get the dataset root directory.
        data_root = cfg.get('data_root', '')
        if not data_root:
            print("Error: data_root not found in config, cannot generate prompt pool")
            print("Error: data_root not found in config, cannot generate prompt pool")
            return
        
        # Resolve class-name and exp-file paths.
        classname_path = model_cfg.get('classname_path', '')
        if not classname_path:
            print("Error: classname_path not found in model config, cannot generate prompt pool")
            return
        
        # Read class names from the class-name file.
        with open(classname_path, 'r') as f:
            class_names = [line.strip() for line in f.readlines() if line.strip()]
        
        # Resolve the validation image directory.
        val_img_dir = os.path.join(data_root, 'img_dir', 'val')
        if not os.path.exists(val_img_dir):
            print(f"Validation image directory does not exist: {val_img_dir}")
            return
        
        # Collect image files for prompt pool generation.
        image_extensions = ['.jpg', '.jpeg', '.png', '.tif', '.tiff']
        image_paths = []
        
        for root, dirs, files in os.walk(val_img_dir):
            for file in files:
                if any(file.lower().endswith(ext) for ext in image_extensions):
                    image_paths.append(os.path.join(root, file))
        
        # Determine how many images to sample for prompt pool construction.
        num_images_to_use = max(1, int(len(image_paths) * pool_building_percent / 100))
        # Use a fixed random seed for reproducible sampling.
        import random
        rng = random.Random(42)  
        sampled_image_paths = rng.sample(image_paths, min(num_images_to_use, len(image_paths)))
        
        print(f"Found {len(image_paths)} total images in dataset")
        print(f"Using {len(sampled_image_paths)} images for prompt pool generation ({pool_building_percent}%)")
        print(f"Classes: {class_names}")
        
        # Generate the prompt pool.
        prompt_pool = generate_prompt_pool(
            image_paths=sampled_image_paths,
            class_names=class_names,
            output_path=prompt_pool_path
        )
        
        print(f"Prompt pool generation completed. Saved to {prompt_pool_path}")
    else:
        print(f"Using existing prompt pool at {prompt_pool_path}")


def build_concept_pool_for_improved_model(cfg, runner):
    """
    Build the concept pool for ImprovedEnhancedSegEarthOV3Segmentation models.
    """
    model_cfg = cfg.get('model', {})
    model_type = model_cfg.get('type', '')
    
    # Check whether this model needs concept-pool construction.
    if 'ImprovedEnhancedSegEarthOV3Segmentation' not in model_type:
        print("Model is not ImprovedEnhancedSegEarthOV3Segmentation, skipping concept pool building.")
        return
    
    # Get the actual model instance, unwrapping DDP when needed.
    model = runner.model
    if hasattr(model, 'module'):
        # Use module to access the wrapped model in distributed mode.
        actual_model = model.module
    else:
        # Use the model directly when it is not wrapped.
        actual_model = model
    
    # Access the test dataset for concept-pool construction.
    test_dataset = runner.test_loop.dataloader.dataset
    
    # Call the model concept-pool builder.
    actual_model.build_concept_pool(test_dataset, seed=42)
    
    print("Concept pool building completed.")


def perform_custom_counting_analysis(cfg, args):
    """
    Run optional custom counting analysis and compare it with SAM3 outputs.
    """
    # Get the model config.
    model_cfg = cfg.get('model', {})
    model_type = model_cfg.get('type', '')
    
    # Check whether counting analysis is enabled.
    perform_counting = model_cfg.get('perform_counting_analysis', False)
    if not perform_counting:
        print("Counting analysis not enabled, skipping.")
        return
    
    print("Starting custom counting analysis...")
    
    # Get the dataset root directory.
    data_root = cfg.get('data_root', '')
    if not data_root:
        print("Error: data_root not found in config, cannot perform counting analysis")
        return
    
    # Get class names.
    class_names = cfg.get('class_names', [])
    if not class_names:
        classname_path = model_cfg.get('classname_path', '')
        if classname_path and os.path.exists(classname_path):
            with open(classname_path, 'r') as f:
                class_names = [line.strip() for line in f.readlines() if line.strip()]
    
    if not class_names:
        print("Error: class_names not found in config, cannot perform counting analysis")
        return
    
    # Resolve the validation image directory.
    val_img_dir = os.path.join(data_root, 'img_dir', 'val')
    if not os.path.exists(val_img_dir):
        print(f"Validation image directory does not exist: {val_img_dir}")
        return
    
    # Define supported image file extensions.
    image_extensions = ['.jpg', '.jpeg', '.png', '.tif', '.tiff']
    image_paths = []
    
    for root, dirs, files in os.walk(val_img_dir):
        for file in files:
            if any(file.lower().endswith(ext) for ext in image_extensions):
                image_paths.append(os.path.join(root, file))
    
    # Limit the counting analysis to the first few images.
    if len(image_paths) > 5:
        image_paths = image_paths[:5]
    
    print(f"Performing counting analysis on {len(image_paths)} images")
    
    # Resolve the expanded prompt pool path for counting analysis.
    prompt_pool_path = model_cfg.get('expanded_prompt_pool_path', None)
    
    # Create the analyzer and run the counting analysis.
    analyzer = CustomCountingAnalyzer()
    results = analyzer.batch_analyze(
        image_paths=image_paths,
        prompt_pool_path=prompt_pool_path,
        class_names=class_names
    )
    
    print("Custom counting analysis completed.")


def main():
    args = parse_args()
    print(os.getcwd())
    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)
    
    # Ensure the expanded prompt pool exists before model initialization.
    check_and_generate_expanded_prompt_pool(cfg, args)
    
    # Ensure the standard prompt pool exists before evaluation.
    check_and_generate_prompt_pool(cfg)
    
    # Add output_dir to the metric config when requested.
    if args.out is not None:
        cfg.test_evaluator['output_dir'] = args.out
        cfg.test_evaluator['keep_results'] = True
    cfg.work_dir = osp.join('./work_dirs',
                            osp.splitext(osp.basename(args.config))[0])

    # trigger_visualization_hook(cfg, args)
    # Initialize the runner after prompt pools are ready.
    runner = Runner.from_cfg(cfg)
    
    # Build the concept pool for models that need it.
    build_concept_pool_for_improved_model(cfg, runner)
    
    # Modify the runner to use only a percentage of data.
    runner = modify_runner_for_percent_data(runner, args.percent, args.seed)
    
    results = runner.test()

    results.update({'Model': cfg.model.model_type,
                    'Dataset': cfg.dataset_type,
                    'Data_Percent': args.percent})

    if runner.rank == 0:
        append_experiment_result('results.xlsx', [results])

    if runner.rank == 0:
        with open(os.path.join(cfg.work_dir, 'results.txt'), 'a') as f:
            f.write(os.path.basename(args.config).split('.')[0] + '\n')
            for k, v in results.items():
                f.write(k + ': ' + str(v) + '\n')
    
    # Run optional custom counting analysis after evaluation completes.
    perform_custom_counting_analysis(cfg, args)


if __name__ == '__main__':
    main()
